"""Import the master price list into dictionary (replace) + catalog_prices (upsert).

Run: uv run python scripts/import_master_pricelist.py [--file catalogs/oriflame_prices_with_calculations_fixed.xlsx]
     uv run python scripts/import_master_pricelist.py --only-missing   (additive, deletes nothing)

Source file shape: one sheet "Прайс-лист" with header columns
Код, Название, ДЦ, ПЦ, Последний каталог. Unlike scripts/import_catalogs.py
(products.json: many catalogs per code) + scripts/import_prices.py (many xlsx
files: full per-catalog price history), this script imports ONE authoritative
recent export where every code carries just its single latest catalog issue
("Последний каталог") and current ДЦ/ПЦ prices.

The two helper tables are treated differently, and the difference is the whole
point (quick task 260902-m9g). `dictionary` is still rebuilt wholesale in one
transaction — that table has one row per code and its own rules.
`catalog_prices` is only UPSERTED: this source owns the (year, number, code)
triples it itself carries and nothing else, so the ~223 000 archive rows that
scripts/import_prices.py imported survive a re-run here. Before this task it
emptied the table too, so running the two importers in either order erased one
of them. `upsert_price_rows` is shared with import_prices.py — ONE ownership
rule with ONE implementation — and its no-None-overwrite rule matters here in
particular: the master price list has no ББ column at all, so every record it
builds carries `points=None`, which must never null the archive's bonus points.

Both tables are pure helper data (D-24), never touching
Product/Batch/Operation/Sale/ledger rows. Dictionary.catalogs becomes a
single-element list (this collapses the prior "history of many catalogs" down
to just the latest one).

Quick task 260902-1d1 added the override-only branch: a справочник code can
exist as a real product and appear in NO price list (the 34 «НЕТ В
СПРАВОЧНИКЕ» codes of the «Офис» inventory). Those codes come from
RUBRIC_OVERRIDES with a name and a rubric but никогда a price, so they land
in dictionary with ``catalogs == []`` and get no catalog_prices row at all.
``--only-missing`` is the non-destructive way to backfill them into a live
database: it inserts the codes that are not there yet and deletes nothing.
"""

import argparse
import json
import re
import sys
from datetime import UTC, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402

from app.config import settings  # noqa: E402
from app.core import new_id, to_cents  # noqa: E402
from app.db import SessionLocal, engine  # noqa: E402
from app.models import CatalogPrice, Dictionary  # noqa: E402
from app.services.backup import create_backup  # noqa: E402
from app.services.catalogs import to_json_code  # noqa: E402
from app.services.rubrics import RUBRIC_OVERRIDES, resolve_name, resolve_rubric  # noqa: E402
from scripts.import_catalogs import export_dictionary  # noqa: E402
from scripts.import_prices import atomic_write, build_price_rows, upsert_price_rows  # noqa: E402

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_FILE = "catalogs/oriflame_prices_with_calculations_fixed.xlsx"
SHEET_NAME = "Прайс-лист"
EXPECTED_HEADERS = ["Код", "Название", "ДЦ", "ПЦ", "Последний каталог"]


class DictionaryReplaceRefused(RuntimeError):
    """The wholesale `dictionary` replace was refused before it deleted anything.

    Mirrors ShadeNameWouldShrink in scripts/import_catalogs.py — the last line
    of defence for the unattended server run. Named `Refused` rather than
    `WouldShrink` because it carries two distinct messages: an empty price list
    (which is not a shrink at all when the table is empty too) and a replace
    that would leave fewer rows stored than there are now.
    """


def parse_last_catalog(value) -> tuple[int, int] | None:
    """"17-2021" -> (2021, 17); handles YYYY-MM, MM-YYYY and YY-MM shapes.

    Mirrors the exact disambiguation heuristic of parse_catalog() in
    scripts/import_prices.py, applied to the "Последний каталог" cell value
    instead of a filename.
    """
    nums = re.findall(r"\d+", str(value).strip())
    if len(nums) < 2:
        return None
    a, b = int(nums[0]), int(nums[1])
    if len(nums[0]) == 4:  # YYYY-MM
        return a, b
    if len(nums[1]) == 4:  # MM-YYYY
        return b, a
    if a > 17:  # YY-MM (a is the year, catalogs never exceed 17)
        return 2000 + a, b
    if b > 17:  # MM-YY
        return 2000 + b, a
    return None


def _cents(value) -> int | None:
    """Positive number -> integer cents; anything else (incl. blank) -> None."""
    if not isinstance(value, (int, float)) or value <= 0:
        return None
    try:
        return to_cents(str(value))
    except ValueError:
        return None


def collect_price_rows(src: Path) -> tuple[dict[str, dict], dict[str, int]]:
    """Read the price-list sheet into {code: row-data} + skip statistics."""
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"Sheet {SHEET_NAME!r} not found in {src} (sheets: {wb.sheetnames})")
    ws = wb[SHEET_NAME]

    header_row = next(ws.iter_rows(values_only=True, max_row=1))
    colmap: dict[str, int] = {}
    for j, cell in enumerate(header_row):
        name = str(cell).strip() if cell is not None else ""
        if name in EXPECTED_HEADERS and name not in colmap:
            colmap[name] = j
    missing = [h for h in EXPECTED_HEADERS if h not in colmap]
    if missing:
        sys.exit(f"Missing expected column(s) {missing} in {src} sheet {SHEET_NAME!r}")

    total_rows = 0
    skipped_missing_code = 0
    skipped_bad_catalog = 0
    collected: dict[str, dict] = {}

    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        total_rows += 1
        code_cell = row[colmap["Код"]] if colmap["Код"] < len(row) else None
        if code_cell is None or str(code_cell).strip() == "":
            skipped_missing_code += 1
            continue
        code = str(code_cell).strip()

        cat = parse_last_catalog(row[colmap["Последний каталог"]] if colmap["Последний каталог"] < len(row) else None)
        if cat is None:
            skipped_bad_catalog += 1
            continue
        year, number = cat

        name_cell = row[colmap["Название"]] if colmap["Название"] < len(row) else None
        name = str(name_cell).strip()[:200] if name_cell is not None and str(name_cell).strip() else None

        consultant_cents = _cents(row[colmap["ДЦ"]] if colmap["ДЦ"] < len(row) else None)
        consumer_cents = _cents(row[colmap["ПЦ"]] if colmap["ПЦ"] < len(row) else None)

        collected[code] = {
            "name": name,
            "year": year,
            "number": number,
            "consumer_cents": consumer_cents,
            "consultant_cents": consultant_cents,
        }
    wb.close()

    stats = {
        "total_rows": total_rows,
        "skipped_missing_code": skipped_missing_code,
        "skipped_bad_catalog": skipped_bad_catalog,
    }
    return collected, stats


def override_only_rows(collected_codes: set[str]) -> dict[str, dict]:
    """RUBRIC_OVERRIDES entries whose code is in no price list.

    These are справочник codes that exist as a real product but appear in no
    price list at all — they get a name and a rubric, but никогда a price.
    """
    return {
        code: entry
        for code, entry in RUBRIC_OVERRIDES.items()
        if code not in collected_codes
    }


def _priceless_dictionary_row(code: str) -> Dictionary:
    """A Dictionary row for an override-only code: name, rubric, no catalog."""
    # resolve_name returns the override name for a code we have one for.
    name = resolve_name(code, code)[:200]
    return Dictionary(
        id=new_id(),
        code=code,
        name=name,
        name_lc=name.lower(),
        catalogs=[],
        rubric=resolve_rubric(code, name),
    )


def build_dictionary_rows(collected: dict[str, dict]) -> list[Dictionary]:
    """One row per price-list code, plus one per override-only code."""
    # CAT-06: assign the rubric and (for bad shade-only names) the corrected
    # full name, both resolved from the code via app.services.rubrics.
    # name_lc is the lowercased shadow column the name filter matches on
    # (LIST-02); set it here or bulk-imported rows are invisible to search.
    dict_rows = []
    for code, data in collected.items():
        name = resolve_name(code, data["name"] or code)[:200]
        dict_rows.append(
            Dictionary(
                id=new_id(),
                code=code,
                name=name,
                name_lc=name.lower(),
                catalogs=[to_json_code(data["year"], data["number"])],
                rubric=resolve_rubric(code, data["name"] or code),
            )
        )
    # 260902-1d1: codes that have an override but no price-list row still
    # belong in the справочник — with an empty catalogs list and no price.
    dict_rows.extend(
        _priceless_dictionary_row(code)
        for code in sorted(override_only_rows(set(collected)))
    )
    return dict_rows


def build_catalog_price_records(collected: dict[str, dict]) -> list[dict]:
    """One 7-key export record per price-list code — the shape upsert_price_rows takes.

    Override-only codes never get one: they have a name and a rubric but
    никогда a price. `points` is always None — the master price list carries no
    ББ column at all, which is exactly the case upsert_price_rows'
    no-None-overwrite rule exists for.
    """
    return [
        {
            "year": data["year"],
            "number": data["number"],
            "code": code,
            "name": data["name"],
            "consumer_cents": data["consumer_cents"],
            "consultant_cents": data["consultant_cents"],
            "points": None,
        }
        for code, data in collected.items()
    ]


def build_catalog_price_rows(collected: dict[str, dict]) -> list[CatalogPrice]:
    """The same records as model objects — one implementation, not two."""
    return build_price_rows(build_catalog_price_records(collected))


def backup_before_replace(engine) -> Path | None:
    """VACUUM INTO snapshot before the wholesale replace — the way back.

    SQLite only: `VACUUM INTO` is a SQLite statement, so on any other dialect
    this prints one skip line and returns None — the policy
    `.env.production.example:25` already states for BACKUP_ON_STARTUP. A server
    run must not crash here, and must not stay silent either.

    The exception `create_backup` raises is deliberately NOT caught: a snapshot
    that cannot be taken must ABORT the import before anything is deleted.
    Takes the engine (not a session) and is called from main() before the
    session is opened on purpose — apply_master_import is called directly by
    unit tests on throwaway engines, and a filesystem side effect inside it
    would spray backups into the developer's real data/backups/.
    """
    if engine.dialect.name != "sqlite":
        print(
            "Rollback snapshot skipped: VACUUM INTO is SQLite-only "
            f"(dialect: {engine.dialect.name})"
        )
        return None
    snapshot = create_backup(engine, Path(settings.backup_dir))
    print(f"Rollback snapshot: {snapshot}")
    return snapshot


def snapshot_dictionary(session) -> Path | None:
    """Dump `dictionary` to a JSON file before the replace — the PORTABLE way back.

    `backup_before_replace` above is `VACUUM INTO`, a SQLite-only statement, so
    on the s1 PostgreSQL deployment (`deploy/DEPLOY.s1.md:26,60`) it prints its
    skip line and takes NOTHING. This snapshot has no dialect at all: it reads
    the table through `import_catalogs.export_dictionary()`, which already dumps
    it in that script's own products.json shape, so the way back is the command
    the shrink refusal already names —
    `scripts/import_catalogs.py --only-missing --file <snapshot>` puts the
    deleted codes back without touching what is there. (Dropping
    `--only-missing` restores every name in the snapshot wholesale, for the case
    where a replace overwrote names rather than deleting rows.)

    Written through `atomic_write`: a half-written rollback artifact would be
    worse than none. An empty table has nothing to lose and writes no file, so a
    clean install does not litter the backup directory. Nothing is caught here —
    like the VACUUM snapshot, no snapshot means no import.
    """
    rows = export_dictionary(session)
    if not rows:
        return None
    stamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    # `dictionary-*.json`, never `myorishop-*.db`: it must not be mistaken for a
    # VACUUM snapshot by prune_backups or by the operator.
    dest = Path(settings.backup_dir) / f"dictionary-{stamp}.json"
    atomic_write(dest, json.dumps(rows, ensure_ascii=False, indent=1) + "\n", newline="\n")
    print(f"Dictionary snapshot ({len(rows)} codes): {dest}")
    print(f"  restore with: scripts/import_catalogs.py --only-missing --file {dest}")
    return dest


def apply_master_import(
    session, collected: dict[str, dict], *, force: bool = False
) -> dict[str, int]:
    """Rebuild `dictionary` wholesale, upsert only this source's price triples.

    Does not commit — the caller owns the transaction. Returns the
    upsert_price_rows stats (inserted / updated / unchanged).

    Both guards live HERE rather than in main(), so every caller is protected.
    THE THRESHOLD IS 0 %, AND IT STAYS 0 % — no tolerance band.
    `deploy/DEPLOY.s1.md:73-121` documents the install order: this importer runs
    FIRST (§4, line 82) and
    `scripts/import_catalogs.py --only-missing --file catalogs/products.json`
    runs after it (§4.1, line 117), so on a clean install `dictionary` is EMPTY
    here — 0 -> 6 856 is growth and the guard is silent on the happy path. The
    case it DOES fire on is a re-run against an already-loaded server, where
    `deploy/DEPLOY.s1.md:101-105` gives the numbers: the master price list
    covers 6 856 codes while the full справочник holds 12 582. That is a 45 %
    loss — precisely the destruction this guard exists to stop, guarded until
    now by nothing but the prose warning at `deploy/DEPLOY.s1.md:94-97`. A 20 %
    tolerance would not even catch that case (45 % > 20 %); it would only weaken
    the small-drift case, so it buys nothing and costs protection.
    """
    if not collected:
        raise DictionaryReplaceRefused(
            "refusing to replace `dictionary` from an empty price list: nothing was "
            "collected, so the replace would delete every row and put nothing back"
        )
    rows = build_dictionary_rows(collected)
    before = session.query(Dictionary).count()
    if len(rows) < before and not force:
        raise DictionaryReplaceRefused(
            f"refusing to replace `dictionary`: stored {before} -> about to write "
            f"{len(rows)}. Restore the fuller справочник with "
            "`scripts/import_catalogs.py --only-missing --file catalogs/products.json`; "
            "pass --force only if this shrink is a deliberate rebuild"
        )
    session.query(Dictionary).delete()
    session.bulk_save_objects(rows)
    return upsert_price_rows(session, build_catalog_price_records(collected))


def insert_missing_dictionary_rows(session, extra: dict[str, dict]) -> list[str]:
    """Additive backfill: insert only the `extra` codes not in dictionary yet.

    Never deletes, never updates an existing row, never writes a CatalogPrice.
    The caller commits. Returns the inserted codes.
    """
    existing = {code for (code,) in session.query(Dictionary.code).all()}
    inserted: list[str] = []
    for code in sorted(extra):
        if code in existing:
            continue
        session.add(_priceless_dictionary_row(code))
        inserted.append(code)
    return inserted


def main() -> None:
    parser = argparse.ArgumentParser(description="Full-replace import of the master price list")
    parser.add_argument("--file", default=DEFAULT_FILE, help="path to the master price list xlsx")
    parser.add_argument(
        "--only-missing",
        action="store_true",
        help="additive: insert only the missing override-only codes; delete nothing",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="permit a full replace that leaves FEWER dictionary rows than are stored",
    )
    args = parser.parse_args()

    # Foot-gun guard, in the style of the sibling scripts: --only-missing deletes
    # nothing at all, so there is no shrink for --force to permit.
    if args.force and args.only_missing:
        sys.exit("--force is meaningless with --only-missing; that mode deletes nothing")

    src = Path(args.file)
    if not src.is_absolute():
        src = PROJECT_ROOT / src
    if not src.is_file():
        sys.exit(f"Source file not found: {src}")

    collected, stats = collect_price_rows(src)
    total_rows = stats["total_rows"]
    skipped_missing_code = stats["skipped_missing_code"]
    skipped_bad_catalog = stats["skipped_bad_catalog"]
    extra = override_only_rows(set(collected))

    # The same empty-input refusal the sibling script has (import_prices.py:988).
    # It guards BOTH modes on purpose: with a degraded parse --only-missing would
    # insert a priceless «Не опознан» row for every override code that actually IS
    # in the price list. It deletes nothing, but it is still junk, and fail-closed
    # is the rule here.
    if not collected:
        sys.exit(
            f"Collected 0 price rows from {src} "
            f"(missing code: {skipped_missing_code}, "
            f"unparsable catalog: {skipped_bad_catalog}) — nothing written"
        )

    if args.only_missing:
        with SessionLocal() as session:
            before_dict = session.query(Dictionary).count()
            inserted = insert_missing_dictionary_rows(session, extra)
            session.commit()
            after_dict = session.query(Dictionary).count()
        print(f"Source: {src}")
        print("Mode: --only-missing (additive, nothing deleted)")
        print(f"Rows read from price list: {len(collected)}")
        print(f"Override-only codes (no price): {len(extra)}")
        print(f"Inserted: {len(inserted)} (skipped, already present: {len(extra) - len(inserted)})")
        print(f"Dictionary: {before_dict} -> {after_dict}")
        return

    # The pre-write half of the summary, and the snapshot, BEFORE the first row
    # is deleted: statistics an operator reads after session.commit() cannot stop
    # anything. The post-write half stays below — those numbers do not exist yet.
    print(f"Source: {src}")
    print(f"Sheet: {SHEET_NAME}")
    print(f"Data rows scanned: {total_rows}")
    print(f"Rows imported: {len(collected)}")
    print(
        "Rows skipped: "
        f"{skipped_missing_code + skipped_bad_catalog} "
        f"(missing code: {skipped_missing_code}, unparsable catalog: {skipped_bad_catalog})"
    )
    print(f"Dictionary rows from overrides only (no price): {len(extra)}")
    backup_before_replace(engine)

    with SessionLocal() as session:
        before_dict = session.query(Dictionary).count()
        before_cp = session.query(CatalogPrice).count()

        # The portable half of the rollback promise: backup_before_replace above
        # is a no-op on PostgreSQL, and s1 IS PostgreSQL. Deliberately here and
        # not inside apply_master_import — that function must stay free of
        # filesystem side effects, because unit tests call it directly.
        snapshot_dictionary(session)

        price_stats = apply_master_import(session, collected, force=args.force)
        session.commit()

        after_dict = session.query(Dictionary).count()
        after_cp = session.query(CatalogPrice).count()
        rubric_filled = (
            session.query(Dictionary).filter(Dictionary.rubric.isnot(None)).count()
        )
        rubric_other = (
            session.query(Dictionary).filter(Dictionary.rubric == "Прочее").count()
        )

    print(f"Dictionary: {before_dict} -> {after_dict} (replaced wholesale)")
    print(f"CatalogPrice: {before_cp} -> {after_cp} (upserted, nothing removed)")
    print(
        f"  inserted: {price_stats['inserted']}  updated: {price_stats['updated']}  "
        f"unchanged: {price_stats['unchanged']}"
    )
    print(f"Rubric assigned: {rubric_filled}/{after_dict} (Прочее: {rubric_other})")


if __name__ == "__main__":
    main()
